import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

class ExcelSafetyChecker(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Excel 안전 검사기")
        self.geometry("700x500")
        self.resizable(False, False)

        # 도착용 표준 항목
        self.standard_headers_arrival = [
            "No", "알림", "발송접수일", "발송접수시간", "구분", "법인", "발송구분", "회수", "미착", "과착", "변상", "운송장번호",
            "고객사주문번호", "인수자타입", "인수자명", "인수완료일시", "접수계정", "발송지", "발송지전화번호", "도착지", "도착지전화번호",
            "보내는분", "보낸분전화번호", "보낸분기타전화번호", "보낸분우편번호", "보낸분주소", "보낸분상세주소", "받는분",
            "도착고객관리", "받는분전화번호", "받는분기타전화번호", "받는분우편번호", "받는분주소", "받는분상세주소", "품목명",
            "배송구분", "포장상태", "수량", "운임구분", "결재구분", "신용", "결재여부", "결재일시", "개별단가", "가로", "세로",
            "높이", "무게", "CBM", "할증운임", "배송운임", "도서운임", "기타운임", "별도운임", "운임합계", "출고비",
            "출고비결재수단", "출고비결재여부", "출고비결재일시", "메모", "노선번호", "발송연선번호", "도착연선번호", "발송지지역",
            "도착지지역", "발송지관할지역", "도착지관할지역", "발송터미널", "도착터미널", "발송터미널하차번호", "도착터미널하차번호"
        ]

        # 발송용 표준 항목
        self.standard_headers_departure = [
            "No", "알림", "발송접수일", "발송접수시간", "구분", "법인", "발송구분", "회수", "미착", "과착", "변상", "운송장번호",
            "고객사주문번호", "인수자타입", "인수자명", "인수완료일시", "접수계정", "발송지", "발송지전화번호", "도착지", "도착지전화번호",
            "보내는분", "고객관리번호", "발송고객관리", "보낸분전화번호", "보낸분기타전화번호", "보낸분우편번호", "보낸분주소",
            "보낸분상세주소", "받는분", "받는분전화번호", "받는분기타전화번호", "받는분우편번호", "받는분주소", "받는분상세주소",
            "품목명", "배송구분", "포장상태", "수량", "운임구분", "결재구분", "신용", "결재여부", "결재일시", "개별단가",
            "가로", "세로", "높이", "무게", "CBM", "할증운임", "배송운임", "도서운임", "기타운임", "별도운임", "운임합계",
            "메모", "노선번호", "발송연선번호", "도착연선번호", "발송지지역", "도착지지역", "발송지관할지역", "도착지관할지역",
            "발송터미널", "도착터미널", "발송터미널하차번호", "도착터미널하차번호", "픽업기사", "발송지수수료", "발송지운임삭제",
            "후불기타운임결재수단", "후불기타운임결재여부", "후불기타운임결재일시", "거래처 체크"
        ]

        # 발송 파일에 반드시 필요한 필드 3개
        self.DEPARTURE_REQUIRED_FIELDS = ["고객관리번호", "발송고객관리", "거래처 체크"]

        # 현재 사용 중인 표준 항목 (버튼 클릭 시 설정)
        self.current_standard_headers = []

        # 누락 항목 저장 변수
        self.missing_headers = []
        # 수정된 셀 정보
        self.modified_cells = []
        # 파일/워크북 관련 변수
        self.file_path = ""
        self.file_type = ""  # 예: ".xlsx" 또는 ".xls"
        self.wb = None      # .xlsx 파일일 때 워크북 객체
        self.modified_dfs = {}

        # 메인 프레임
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 업로드 영역
        upload_frame = ttk.LabelFrame(main_frame, text="엑셀 파일 업로드")
        upload_frame.pack(fill=tk.X, pady=5)

        # 도착 파일 선택 버튼
        arrival_button = ttk.Button(upload_frame, text="도착 파일 선택", command=self.select_arrival_file)
        arrival_button.pack(side=tk.LEFT, padx=10, pady=5)

        # 발송 파일 선택 버튼
        departure_button = ttk.Button(upload_frame, text="발송 파일 선택", command=self.select_departure_file)
        departure_button.pack(side=tk.LEFT, padx=10, pady=5)

        self.file_label = ttk.Label(upload_frame, text="선택된 파일: 없음", foreground="blue")
        self.file_label.pack(padx=10, pady=(0, 10), anchor=tk.W)

        # 상세보기 버튼
        self.details_visible = False
        self.toggle_details_btn = ttk.Button(main_frame, text="상세", command=self.toggle_details)
        self.toggle_details_btn.pack(padx=10, pady=5, anchor=tk.W)

        # 상세 영역
        self.details_frame = ttk.Frame(main_frame)

        # 표준 샘플 항목 등록 영역
        register_frame = ttk.LabelFrame(self.details_frame, text="표준 샘플 항목 등록")
        register_frame.pack(fill=tk.X, pady=5)
        register_label = ttk.Label(register_frame, text="현재 적용되는 표준 항목(헤더):")
        register_label.pack(padx=10, pady=2, anchor=tk.W)
        self.standard_text = tk.Text(register_frame, height=4)
        self.standard_text.pack(padx=10, pady=2, fill=tk.X)
        self.standard_text.config(state="disabled")

        # 수정된 셀 미리보기 영역
        preview_frame = ttk.LabelFrame(self.details_frame, text="수정된 셀 미리보기")
        preview_frame.pack(fill=tk.BOTH, pady=5, expand=True)
        preview_columns = ("cell", "old_value", "new_value")
        self.tree = ttk.Treeview(preview_frame, columns=preview_columns, show="headings", height=10)
        self.tree.heading("cell", text="셀 위치")
        self.tree.heading("old_value", text="변경 전 값")
        self.tree.heading("new_value", text="변경 후 값")
        self.tree.column("cell", width=120)
        self.tree.column("old_value", width=400)
        self.tree.column("new_value", width=400)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        preview_scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=preview_scrollbar.set)
        preview_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 표준 항목 비교 영역
        compare_frame = ttk.LabelFrame(self.details_frame, text="표준 항목 비교")
        compare_frame.pack(fill=tk.BOTH, pady=5, expand=True)
        compare_columns = ("index", "expected", "actual", "status")
        self.compare_tree = ttk.Treeview(compare_frame, columns=compare_columns, show="headings", height=8)
        self.compare_tree.heading("index", text="순번")
        self.compare_tree.heading("expected", text="표준 항목")
        self.compare_tree.heading("actual", text="실제 항목")
        self.compare_tree.heading("status", text="비고")
        self.compare_tree.column("index", width=50)
        self.compare_tree.column("expected", width=200)
        self.compare_tree.column("actual", width=200)
        self.compare_tree.column("status", width=100)
        self.compare_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        compare_scrollbar = ttk.Scrollbar(compare_frame, orient="vertical", command=self.compare_tree.yview)
        self.compare_tree.configure(yscroll=compare_scrollbar.set)
        compare_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 처리 결과 라벨
        self.result_label = ttk.Label(main_frame, text="처리 결과: -", foreground="green", font=("맑은 고딕", 10, "bold"))
        self.result_label.pack(pady=5)

    # 도착 파일 선택 버튼 콜백
    def select_arrival_file(self):
        self.current_standard_headers = self.standard_headers_arrival
        self.refresh_standard_text()
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")],
            title="도착 엑셀 파일 선택"
        )
        if file_path:
            self.process_file(file_path, mode="arrival")

    # 발송 파일 선택 버튼 콜백
    def select_departure_file(self):
        self.current_standard_headers = self.standard_headers_departure
        self.refresh_standard_text()
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")],
            title="발송 엑셀 파일 선택"
        )
        if file_path:
            self.process_file(file_path, mode="departure")

    # 현재 표준 항목(헤더)을 Text 위젯에 표시
    def refresh_standard_text(self):
        self.standard_text.config(state="normal")
        self.standard_text.delete("1.0", tk.END)
        self.standard_text.insert("1.0", "\n".join(self.current_standard_headers))
        self.standard_text.config(state="disabled")

    # 상세보기 토글
    def toggle_details(self):
        if self.details_visible:
            self.details_frame.pack_forget()
            self.details_visible = False
            self.toggle_details_btn.config(text="상세")
        else:
            self.details_frame.pack(fill=tk.BOTH, expand=True)
            self.details_visible = True
            self.toggle_details_btn.config(text="상세 숨기기")

    # 파일 처리
    def process_file(self, file_path, mode):
        self.file_path = file_path
        self.file_label.config(text=f"선택된 파일: {os.path.basename(file_path)}", foreground="blue")
        self.modified_cells.clear()
        self.missing_headers = []
        total_sheets = 0
        total_cells_checked = 0
        modified_count = 0

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in [".xlsx", ".xls"]:
            messagebox.showwarning("경고", "지원하지 않는 파일 형식입니다. (.xlsx 또는 .xls)")
            return

        # 파일 확장자를 저장 (점 포함)
        self.file_type = ext

        actual_headers = []

        # 엑셀 파일 로드
        if ext == ".xlsx":
            try:
                wb = openpyxl.load_workbook(file_path, data_only=False)
                total_sheets = len(wb.worksheets)
                for ws in wb.worksheets:
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            total_cells_checked += 1
                            # 셀의 값이 문자열이고 "="로 시작하거나, 셀의 데이터 타입이 'f'(formula)인 경우 처리
                            if cell.value and (isinstance(cell.value, str) and cell.value.startswith("=") or cell.data_type == "f"):
                                old_value = cell.value
                                if isinstance(cell.value, str) and cell.value.startswith("="):
                                    new_value = cell.value[1:]  # = 문자만 제거
                                else:
                                    new_value = cell.value  # 수식인 경우 원래 값 유지
                                cell.value = new_value  # str() 변환 제거
                                modified_count += 1
                                self.modified_cells.append((cell.coordinate, old_value, new_value))
                self.wb = wb
                # 첫 시트 헤더
                first_ws = wb.worksheets[0]
                actual_headers = [cell.value if cell.value is not None else "" for cell in first_ws[1]]
            except Exception as e:
                messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다.\n{str(e)}")
                return

        elif ext == ".xls":
            try:
                sheets_dict = pd.read_excel(file_path, sheet_name=None, header=None, engine="xlrd")
                total_sheets = len(sheets_dict)
                self.modified_dfs = {}
                for sheet_name, df in sheets_dict.items():
                    if df.empty or df.shape[0] < 2:
                        self.modified_dfs[sheet_name] = df
                        continue
                    df_modified = df.copy()
                    for row in range(1, df_modified.shape[0]):
                        for col in range(df_modified.shape[1]):
                            total_cells_checked += 1
                            cell_value = df_modified.iat[row, col]
                            if isinstance(cell_value, str) and cell_value.startswith("="):
                                old_value = cell_value
                                if isinstance(cell_value, str) and cell_value.startswith("="):
                                    new_value = cell_value[1:]  # = 문자만 제거
                                else:
                                    new_value = cell_value  # 수식인 경우 원래 값 유지
                                df_modified.iat[row, col] = new_value
                                modified_count += 1
                                cell_coordinate = f"{get_column_letter(col+1)}{row+1}"
                                self.modified_cells.append((cell_coordinate, old_value, new_value))
                    self.modified_dfs[sheet_name] = df_modified

                if sheets_dict:
                    first_sheet = list(sheets_dict.keys())[0]
                    df_first = sheets_dict[first_sheet]
                    if not df_first.empty:
                        actual_headers = list(df_first.iloc[0])
                        actual_headers = [str(item) if pd.notna(item) else "" for item in actual_headers]
            except Exception as e:
                messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다.\n{str(e)}")
                return

        # 모드에 따른 헤더 검증
        if mode == "arrival":
            # 도착 모드: 발송용 필드가 있으면 경고
            for field in self.DEPARTURE_REQUIRED_FIELDS:
                if field in actual_headers:
                    messagebox.showwarning("파일 오류", f"이건 발송 파일입니다. 도착 파일이 아닙니다. (필드: {field} 발견)")
                    return
        elif mode == "departure":
            # 발송 모드: 발송 필수 필드가 누락되면 경고
            for field in self.DEPARTURE_REQUIRED_FIELDS:
                if field not in actual_headers:
                    messagebox.showwarning("파일 오류", f"이건 도착 파일입니다. 발송 파일이 아닙니다. (필드: {field} 누락)")
                    return

        # 표준 항목 비교
        if self.current_standard_headers and actual_headers:
            self.compare_headers(actual_headers)

        self.result_label.config(
            text=f"총 시트 수: {total_sheets}개, 검사한 셀 수: {total_cells_checked}개, 수정된 셀 수: {modified_count}개"
        )
        self.update_treeview()

        if modified_count == 0:
            messagebox.showinfo("알림", "이상없슴")
            return

        if self.missing_headers:
            missing_str = ", ".join(self.missing_headers)
            messagebox.showwarning("누락 경고", f"다음 표준 항목이 누락되었습니다: {missing_str}\n엑셀 파일을 다시 다운로드 하십시오.")
            return

        self.save_file()

    # 수정된 셀 미리보기 트리뷰 업데이트
    def update_treeview(self):
        for row_id in self.tree.get_children():
            self.tree.delete(row_id)
        for item in self.modified_cells:
            self.tree.insert("", tk.END, values=item)

    # 표준 항목 비교
    def compare_headers(self, actual_headers):
        comparison = []
        missing_headers = []
        for idx, item in enumerate(self.current_standard_headers, start=1):
            if item not in actual_headers:
                comparison.append((idx, item, "", "누락"))
                missing_headers.append(item)
        for idx, item in enumerate(actual_headers, start=1):
            if item not in self.current_standard_headers:
                comparison.append((idx, "", item, "추가"))
        for row_id in self.compare_tree.get_children():
            self.compare_tree.delete(row_id)
        for row_data in comparison:
            self.compare_tree.insert("", tk.END, values=row_data)
        self.missing_headers = missing_headers

    # 파일 저장
    def save_file(self):
        if not self.file_path:
            return
        initial_name = os.path.basename(self.file_path)
        save_path = filedialog.asksaveasfilename(
            initialfile=initial_name,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="수정된 엑셀 파일 저장"
        )
        if save_path:
            try:
                if self.file_type == ".xlsx":
                    self.wb.save(save_path)
                elif self.file_type == ".xls":
                    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                        for sheet_name, df in self.modified_dfs.items():
                            df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
                messagebox.showinfo("완료", f"수정된 파일이 저장되었습니다.\n{save_path}")
            except Exception as e:
                messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다.\n{str(e)}")


if __name__ == "__main__":
    app = ExcelSafetyChecker()
    app.mainloop()
